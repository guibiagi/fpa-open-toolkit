"""Tests for xlsx_report.py — validates the 6-sheet FP&A Excel report."""

from __future__ import annotations

import datetime
import os
import tempfile

import pandas as pd
import pytest
from openpyxl import load_workbook

# Import after ensuring the project is on sys.path
from data_generation.synthetic_generator import generate_all
from export.xlsx_report import XLSXReport, _format_brl, _format_date, generate_report
from kpis.financial_kpis import FinancialKPIs


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def datasets():
    """Generate synthetic datasets once for all tests."""
    return generate_all(seed=42)


@pytest.fixture(scope="module")
def kpis(datasets):
    """Create KPIs from datasets."""
    return FinancialKPIs(datasets)


@pytest.fixture
def temp_path():
    """Temporary xlsx path that cleans up after test."""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    yield path
    if os.path.exists(path):
        os.unlink(path)


# ---------------------------------------------------------------------------
# Format helpers
# ---------------------------------------------------------------------------


class TestFormatHelpers:
    """Unit tests for formatting utilities."""

    def test_format_brl_positive(self):
        assert _format_brl(1234567.89) == "1.234.567,89"

    def test_format_brl_round_number(self):
        assert _format_brl(1000) == "1.000,00"

    def test_format_brl_zero(self):
        assert _format_brl(0) == "0,00"

    def test_format_brl_negative(self):
        assert _format_brl(-500.5) == "-500,50"

    def test_format_brl_none(self):
        assert _format_brl(None) == "0,00"

    def test_format_date_python(self):
        dt = datetime.date(2026, 5, 26)
        assert _format_date(dt) == "26/05/2026"

    def test_format_date_pandas(self):
        ts = pd.Timestamp("2026-05-26")
        assert _format_date(ts) == "26/05/2026"

    def test_format_date_string_fallback(self):
        assert _format_date("hello") == "hello"


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------


class TestXLSXReport:
    """Integration tests — generate report and validate structure."""

    def test_save_creates_file(self, datasets, temp_path):
        """Report.save() creates a valid .xlsx file."""
        report = XLSXReport(datasets=datasets)
        result = report.save(temp_path)
        assert result == temp_path
        assert os.path.exists(temp_path)
        assert os.path.getsize(temp_path) > 1000  # non-trivial

    def test_generate_report_convenience(self, datasets, temp_path):
        """Convenience function works."""
        path = generate_report(datasets=datasets, output_path=temp_path)
        assert path == temp_path
        assert os.path.exists(temp_path)

    def test_report_has_six_sheets(self, datasets, temp_path):
        """Report must have exactly 6 sheets as per spec."""
        report = XLSXReport(datasets=datasets)
        report.save(temp_path)

        wb = load_workbook(temp_path)
        sheet_names = wb.sheetnames

        assert len(sheet_names) == 6
        assert "Resumo Executivo" in sheet_names
        assert "Forecast de Receita" in sheet_names
        assert "Fluxo de Caixa" in sheet_names
        assert "Capital de Giro" in sheet_names
        assert "Endividamento" in sheet_names
        assert "Dados Brutos" in sheet_names

    def test_resumo_has_title(self, datasets, temp_path):
        """Resumo Executivo sheet has title."""
        report = XLSXReport(datasets=datasets)
        report.save(temp_path)

        wb = load_workbook(temp_path)
        ws = wb["Resumo Executivo"]
        assert "FP&A Open Toolkit" in str(ws.cell(row=1, column=1).value)

    def test_resumo_has_date_stamp(self, datasets, temp_path):
        """Resumo Executivo has generation date."""
        report = XLSXReport(datasets=datasets)
        report.save(temp_path)

        wb = load_workbook(temp_path)
        ws = wb["Resumo Executivo"]
        # Find the "Gerado em:" cell somewhere in the sheet
        found = False
        for row in ws.iter_rows(max_col=6, max_row=15):
            for cell in row:
                if cell.value and "Gerado em:" in str(cell.value):
                    found = True
                    break
        assert found, "Date stamp 'Gerado em:' not found in Resumo sheet"

    def test_headers_are_bold(self, datasets, temp_path):
        """All header rows use bold font."""
        report = XLSXReport(datasets=datasets)
        report.save(temp_path)

        wb = load_workbook(temp_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 1:
                continue
            # First row is title; check for header rows (bold + filled)
            bold_cells = 0
            for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row)):
                for cell in row:
                    if cell.font and cell.font.bold and cell.fill and cell.fill.start_color:
                        bold_cells += 1
            # At least one header row should exist per sheet
            assert bold_cells > 0, f"Sheet '{sheet_name}' has no bold+filled headers"


class TestXLSXReportWithKPIs:
    """Tests with KPIs and forecast/cashflow data."""

    @pytest.fixture(scope="class")
    def datasets(self):
        return generate_all(seed=42)

    @pytest.fixture(scope="class")
    def kpis(self, datasets):
        return FinancialKPIs(datasets)

    def test_report_with_kpis(self, datasets, kpis):
        """Report with KPIs generates successfully."""
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        try:
            report = XLSXReport(datasets=datasets, kpis=kpis)
            report.save(path)

            wb = load_workbook(path)

            # Capital de Giro should have NCG data
            ws = wb["Capital de Giro"]
            assert ws.cell(row=1, column=1).value is not None

            # Endividamento should have debt data
            ws_debt = wb["Endividamento"]
            assert ws_debt.cell(row=1, column=1).value is not None

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_working_capital_has_ncg(self, datasets, kpis):
        """Capital de Giro sheet contains NCG value."""
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        try:
            report = XLSXReport(datasets=datasets, kpis=kpis)
            report.save(path)

            wb = load_workbook(path)
            ws = wb["Capital de Giro"]

            # Find "Necessidade de Capital de Giro" text
            found_ncg = False
            for row in ws.iter_rows(max_col=2, max_row=15):
                for cell in row:
                    if cell.value and "Capital de Giro" in str(cell.value):
                        found_ncg = True
                        break
            assert found_ncg, "NCG label not found in Capital de Giro sheet"

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_debt_has_metrics(self, datasets, kpis):
        """Endividamento sheet contains debt metrics."""
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        try:
            report = XLSXReport(datasets=datasets, kpis=kpis)
            report.save(path)

            wb = load_workbook(path)
            ws = wb["Endividamento"]

            # Find "Dívida Total" text
            found = False
            for row in ws.iter_rows(max_col=2, max_row=20):
                for cell in row:
                    if cell.value and "Dívida Total" in str(cell.value):
                        found = True
                        break
            assert found, "Dívida Total not found in Endividamento sheet"

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_raw_data_has_all_sections(self, datasets, kpis):
        """Dados Brutos has all 6 dataset sections."""
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        try:
            report = XLSXReport(datasets=datasets, kpis=kpis)
            report.save(path)

            wb = load_workbook(path)
            ws = wb["Dados Brutos"]

            expected_labels = [
                "Faturamento Histórico",
                "Contas a Receber",
                "Contas a Pagar",
                "Estoque",
                "Dívida",
                "Custo de Vendas",
            ]

            for label in expected_labels:
                found = False
                for row in ws.iter_rows(max_col=1, max_row=ws.max_row):
                    for cell in row:
                        if cell.value and label in str(cell.value):
                            found = True
                            break
                assert found, f"'{label}' section not found in Dados Brutos sheet"

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_empty_forecast_handled_gracefully(self, datasets, kpis):
        """Report with empty forecast/cashflow should not crash."""
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        try:
            report = XLSXReport(datasets=datasets, kpis=kpis)  # no forecast/cashflow
            report.save(path)

            wb = load_workbook(path)

            # Forecast sheet should show a message
            ws = wb["Forecast de Receita"]
            found = False
            for row in ws.iter_rows(max_col=1, max_row=10):
                for cell in row:
                    if cell.value and "não disponível" in str(cell.value):
                        found = True
                        break
            assert found, "Forecast sheet should warn about missing data"

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_report_without_kpis_works(self, datasets):
        """Report without KPIs should not crash — just show warnings."""
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        try:
            report = XLSXReport(datasets=datasets, kpis=None)
            report.save(path)

            wb = load_workbook(path)
            assert len(wb.sheetnames) == 6
            assert os.path.getsize(path) > 1000

        finally:
            if os.path.exists(path):
                os.unlink(path)
