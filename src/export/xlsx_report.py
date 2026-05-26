"""Excel report generator for FP&A Open Toolkit.

Generates a 6-sheet consolidated .xlsx report with Brazilian formatting:
- Bold headers, auto-fitted column widths
- BRL currency (R$ X.XXX,XX)
- Dates in DD/MM/AAAA
- All engines and base datasets included

Implements spec v1.1 §7.5 — exportação de relatórios.

Usage::

    from export.xlsx_report import XLSXReport

    report = XLSXReport(
        datasets=generate_all(),
        forecast_df=forecast_df,
        cashflow_df=cashflow_df,
        kpis=FinancialKPIs(datasets),
    )
    report.save("data/outputs/relatorio_fpa.xlsx")
"""

from __future__ import annotations

import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGNMENT = Alignment(vertical="center")
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")
DATE_ALIGNMENT = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="333333")

BRL_FORMAT = '#.##0,00'
PCT_FORMAT = '0.0%'

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _format_date(dt: Any) -> str:
    """Convert datetime to DD/MM/AAAA string."""
    if dt is None:
        return ""
    if pd.isna(dt):
        return ""
    if isinstance(dt, pd.Timestamp):
        return dt.strftime("%d/%m/%Y")
    if isinstance(dt, (datetime.date, datetime.datetime)):
        return dt.strftime("%d/%m/%Y")
    return str(dt)


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)
        adjusted = max(min(max_len + 2, max_width), min_width)
        ws.column_dimensions[col_letter].width = adjusted


def _style_header_row(ws, ncols: int, row: int = 1) -> None:
    """Apply header styling to a row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data_cell(ws, row: int, col: int, is_date: bool = False, is_number: bool = False) -> None:
    """Apply data styling to a single cell."""
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if is_date:
        cell.alignment = DATE_ALIGNMENT
    elif is_number:
        cell.alignment = NUMBER_ALIGNMENT
    else:
        cell.alignment = DATA_ALIGNMENT


def _write_dataframe(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
    date_cols: Optional[List[str]] = None,
    number_cols: Optional[List[str]] = None,
) -> int:
    """Write a DataFrame to a worksheet with full styling.

    Returns the row number after the last data row.
    """
    if date_cols is None:
        date_cols = []
    if number_cols is None:
        number_cols = []

    # Map column names to 1-based indices
    cols = list(df.columns)
    date_col_indices = {cols.index(c) + 1 for c in date_cols if c in cols}
    number_col_indices = {cols.index(c) + 1 for c in number_cols if c in cols}

    # Headers
    for j, col_name in enumerate(cols):
        cell = ws.cell(row=start_row, column=start_col + j, value=col_name)
    _style_header_row(ws, len(cols), row=start_row)

    # Data rows
    for i, (_, row_data) in enumerate(df.iterrows()):
        row_num = start_row + 1 + i
        for j, col_name in enumerate(cols):
            val = row_data[col_name]
            col_num = start_col + j

            # Convert to native Python types
            if isinstance(val, (pd.Timestamp, datetime.datetime)):
                ws.cell(row=row_num, column=col_num, value=_format_date(val))
                _style_data_cell(ws, row_num, col_num, is_date=True)
            elif isinstance(val, float) and pd.isna(val):
                ws.cell(row=row_num, column=col_num, value="")
                _style_data_cell(ws, row_num, col_num)
            elif isinstance(val, (int, float, np.integer, np.floating)):
                ws.cell(row=row_num, column=col_num, value=float(val))
                is_num = (col_num - start_col + 1) in number_col_indices
                _style_data_cell(ws, row_num, col_num, is_number=is_num)
            else:
                ws.cell(row=row_num, column=col_num, value=str(val))
                _style_data_cell(ws, row_num, col_num)

    return start_row + len(df) + 1


def _write_title(ws, row: int, col: int, text: str) -> int:
    """Write a title and return next available row."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 5)
    return row + 2


def _write_kpi_card(ws, row: int, col: int, label: str, value: str, value_col: int = 3) -> int:
    """Write a KPI label-value pair. Returns next card start row."""
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = Font(name="Calibri", size=9, color="666666")

    value_cell = ws.cell(row=row + 1, column=col, value=value)
    value_cell.font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    value_cell.alignment = Alignment(horizontal="left")

    return row + 3


import numpy as np


# ---------------------------------------------------------------------------
# Main report class
# ---------------------------------------------------------------------------


class XLSXReport:
    """Generate a 6-sheet consolidated FP&A Excel report."""

    def __init__(
        self,
        datasets: dict[str, pd.DataFrame],
        forecast_df: Optional[pd.DataFrame] = None,
        cashflow_df: Optional[pd.DataFrame] = None,
        kpis: Optional[Any] = None,
    ) -> None:
        self.datasets = datasets
        self.forecast_df = forecast_df or pd.DataFrame()
        self.cashflow_df = cashflow_df or pd.DataFrame()
        self.kpis = kpis

    # ── Sheet 1: Resumo Executivo ──────────────────────────────────────

    def _write_resumo(self, wb) -> None:
        ws = wb.active
        ws.title = "Resumo Executivo"

        row = 1
        row = _write_title(ws, row, 1, "FP&A Open Toolkit — Resumo Executivo")

        # Date stamp
        stamp_cell = ws.cell(row=row, column=1, value=f"Gerado em: {_format_date(datetime.date.today())}")
        stamp_cell.font = Font(name="Calibri", size=9, color="999999")
        row += 2

        # KPI Cards — 2 rows of 4
        metrics = self._collect_metrics()

        card_labels = [
            ("Receita 12M", "R$"),
            ("Crescimento YoY", ""),
            ("Saldo Caixa Proj.", "R$"),
            ("Menor Saldo Proj.", "R$"),
            ("Estoque Atual", "R$"),
            ("Dívida Total", "R$"),
            ("NCG", "R$"),
            ("Ciclo Financeiro", ""),
        ]

        for i, (label, prefix) in enumerate(card_labels):
            r = row + (i // 4) * 3
            c = 1 + (i % 4) * 2

            val = metrics.get(label, 0.0)

            if prefix == "R$" and isinstance(val, (int, float)):
                display = f"R$ {_format_brl(val)}"
            elif label == "Crescimento YoY" and isinstance(val, (int, float)):
                display = _format_brl(val * 100) + "%"  # fraction → percentage
            elif label == "Ciclo Financeiro" and isinstance(val, (int, float)):
                display = f"{val:.0f} dias"
            elif isinstance(val, str):
                display = val
            else:
                display = str(val)

            _write_kpi_card(ws, r, c, label, display)

        row = max(row + 6, 16)

        # Revenue history table
        ws.cell(row=row, column=1, value="Receita Mensal Histórica").font = SUBTITLE_FONT
        row += 1

        fat = self.datasets.get("faturamento_historico", pd.DataFrame())
        if not fat.empty:
            monthly = (
                fat.groupby("data_mes")["faturamento"]
                .sum()
                .reset_index()
                .sort_values("data_mes")
                .tail(12)
            )
            monthly["data_mes"] = pd.to_datetime(monthly["data_mes"])
            monthly.columns = ["Mês", "Faturamento (R$)"]
            row = _write_dataframe(ws, monthly, start_row=row, number_cols=["Faturamento (R$)"]) + 1

        # Forecast summary
        if not self.forecast_df.empty:
            ws.cell(row=row, column=1, value="Resumo do Forecast (Próximos 6 Meses)").font = SUBTITLE_FONT
            row += 1
            fc = self.forecast_df.copy()
            fc["data_mes"] = pd.to_datetime(fc["data_mes"])
            fc = fc.head(6)
            fc.columns = ["Mês", "Base (R$)", "Otimista (R$)", "Pessimista (R$)"]
            row = _write_dataframe(
                ws, fc, start_row=row,
                number_cols=["Base (R$)", "Otimista (R$)", "Pessimista (R$)"],
            )

        _auto_width(ws)

    def _collect_metrics(self) -> dict[str, Any]:
        """Collect all KPI values from the KPIs engine or compute from data."""
        if self.kpis:
            k = self.kpis
            return {
                "Receita 12M": k.receita_ultimos_12m(),
                "Crescimento YoY": k.crescimento_yoy(),
                "Saldo Caixa Proj.": k.saldo_final_projetado(),
                "Menor Saldo Proj.": k.menor_saldo_projetado(),
                "Estoque Atual": k.estoque_total(),
                "Dívida Total": k.divida_total(),
                "NCG": k.ncg(),
                "Ciclo Financeiro": k.ciclo_financeiro(),
            }

        # Fallback: compute from raw data
        fat = self.datasets.get("faturamento_historico", pd.DataFrame())
        if fat.empty:
            return {}

        fat["data_mes"] = pd.to_datetime(fat["data_mes"])
        monthly = fat.groupby("data_mes")["faturamento"].sum().sort_index()
        receita_12m = float(monthly.tail(12).sum())

        crescimento_yoy = 0.0
        if len(monthly) >= 24:
            recent = monthly.iloc[-12:].sum()
            prior = monthly.iloc[-24:-12].sum()
            if prior > 0:
                crescimento_yoy = float((recent - prior) / prior)

        return {
            "Receita 12M": receita_12m,
            "Crescimento YoY": crescimento_yoy,
            "Saldo Caixa Proj.": 0.0,
            "Menor Saldo Proj.": 0.0,
            "Estoque Atual": 0.0,
            "Dívida Total": 0.0,
            "NCG": 0.0,
            "Ciclo Financeiro": 0.0,
        }

    # ── Sheet 2: Forecast de Receita ───────────────────────────────────

    def _write_forecast(self, wb) -> None:
        ws = wb.create_sheet("Forecast de Receita")
        row = 1
        row = _write_title(ws, row, 1, "Forecast de Receita")

        ws.cell(row=row, column=1, value="Projeção de faturamento — modelo: média móvel 3M + tendência + sazonalidade").font = Font(name="Calibri", size=9, color="999999")
        row += 2

        if self.forecast_df.empty:
            ws.cell(row=row, column=1, value="Forecast não disponível. Execute revenue_forecast primeiro.").font = DATA_FONT
        else:
            fc = self.forecast_df.copy()
            fc["data_mes"] = pd.to_datetime(fc["data_mes"])
            cols_map = {
                "data_mes": "Mês",
                "forecast_base": "Cenário Base (R$)",
                "forecast_otimista": "Cenário Otimista (R$)",
                "forecast_pessimista": "Cenário Pessimista (R$)",
            }
            fc = fc[list(cols_map.keys())].rename(columns=cols_map)
            _write_dataframe(
                ws, fc, start_row=row,
                number_cols=["Cenário Base (R$)", "Cenário Otimista (R$)", "Cenário Pessimista (R$)"],
            )

        _auto_width(ws)

    # ── Sheet 3: Fluxo de Caixa Projetado ──────────────────────────────

    def _write_cashflow(self, wb) -> None:
        ws = wb.create_sheet("Fluxo de Caixa")
        row = 1
        row = _write_title(ws, row, 1, "Fluxo de Caixa Projetado — 90 Dias")

        if self.cashflow_df.empty:
            ws.cell(row=row, column=1, value="Fluxo de caixa não disponível. Execute cashflow_projection primeiro.").font = DATA_FONT
        else:
            cf = self.cashflow_df.copy()
            cf["data"] = pd.to_datetime(cf["data"])
            cols_map = {
                "data": "Data",
                "entradas_previstas": "Entradas (R$)",
                "saidas_previstas": "Saídas (R$)",
                "saldo_inicial_dia": "Saldo Inicial (R$)",
                "saldo_final_dia": "Saldo Final (R$)",
                "observacao": "Observação",
            }
            available = [k for k in cols_map if k in cf.columns]
            cf = cf[available].rename(columns={k: v for k, v in cols_map.items() if k in available})
            _write_dataframe(
                ws, cf, start_row=row,
                number_cols=["Entradas (R$)", "Saídas (R$)", "Saldo Inicial (R$)", "Saldo Final (R$)"],
            )

        _auto_width(ws)

    # ── Sheet 4: Capital de Giro ───────────────────────────────────────

    def _write_working_capital(self, wb) -> None:
        ws = wb.create_sheet("Capital de Giro")
        row = 1
        row = _write_title(ws, row, 1, "Análise de Capital de Giro")

        if self.kpis:
            nc_data = {
                "Indicador": [
                    "Contas a Receber (Aberto + Atrasado)",
                    "Contas a Pagar (Aberto)",
                    "Estoque Total",
                    "Necessidade de Capital de Giro (NCG)",
                ],
                "Valor (R$)": [
                    self.kpis.contas_receber_saldo(),
                    self.kpis.contas_pagar_saldo(),
                    self.kpis.estoque_total(),
                    self.kpis.ncg(),
                ],
            }
            nc_df = pd.DataFrame(nc_data)
            _write_dataframe(ws, nc_df, start_row=row, number_cols=["Valor (R$)"])
            row = 6

            # Financial cycle
            ciclo_data = {
                "Indicador": [
                    "Prazo Médio de Recebimento (PMR)",
                    "Prazo Médio de Estoque (PME)",
                    "Prazo Médio de Pagamento (PMP)",
                    "Ciclo Financeiro (PMR+PME-PMP)",
                ],
                "Dias": [
                    self.kpis.prazo_medio_recebimento(),
                    self.kpis.prazo_medio_estoque(),
                    self.kpis.prazo_medio_pagamento(),
                    self.kpis.ciclo_financeiro(),
                ],
            }
            ciclo_df = pd.DataFrame(ciclo_data)
            ws.cell(row=row, column=1, value="Ciclo Financeiro").font = SUBTITLE_FONT
            _write_dataframe(ws, ciclo_df, start_row=row + 1, number_cols=["Dias"])
        else:
            ws.cell(row=row, column=1, value="KPIs não disponíveis. Execute financial_kpis primeiro.").font = DATA_FONT

        _auto_width(ws)

    # ── Sheet 5: Endividamento ─────────────────────────────────────────

    def _write_debt(self, wb) -> None:
        ws = wb.create_sheet("Endividamento")
        row = 1
        row = _write_title(ws, row, 1, "Análise de Endividamento")

        if self.kpis:
            debt_data = {
                "Indicador": [
                    "Dívida Total",
                    "Juros Mensais",
                    "Dívida / Receita Anualizada",
                    "Cobertura de Juros (EBITDA est.)",
                ],
                "Valor": [
                    f"R$ {_format_brl(self.kpis.divida_total())}",
                    f"R$ {_format_brl(self.kpis.juros_mensais())}",
                    _format_brl(self.kpis.divida_receita_anualizada() * 100) + "%",
                    f"{self.kpis.cobertura_juros():.1f}x",
                ],
            }
            debt_df = pd.DataFrame(debt_data)
            _write_dataframe(ws, debt_df, start_row=row)
            row = 6

            # Debt by type
            by_type = self.kpis.divida_por_tipo()
            if by_type:
                ws.cell(row=row, column=1, value="Dívida por Tipo").font = SUBTITLE_FONT
                type_data = {"Tipo": list(by_type.keys()), "Saldo Devedor (R$)": list(by_type.values())}
                type_df = pd.DataFrame(type_data)
                _write_dataframe(ws, type_df, start_row=row + 1, number_cols=["Saldo Devedor (R$)"])
                row += 3 + len(by_type)

            # Debt evolution
            divida = self.datasets.get("divida", pd.DataFrame())
            if not divida.empty:
                ws.cell(row=row, column=1, value="Evolução Mensal da Dívida").font = SUBTITLE_FONT
                evo = (
                    divida.groupby("data_mes")["saldo_devedor"]
                    .sum()
                    .reset_index()
                    .sort_values("data_mes")
                )
                evo["data_mes"] = pd.to_datetime(evo["data_mes"])
                evo.columns = ["Mês", "Saldo Devedor (R$)"]
                _write_dataframe(ws, evo, start_row=row + 1, number_cols=["Saldo Devedor (R$)"])
        else:
            ws.cell(row=row, column=1, value="KPIs não disponíveis. Execute financial_kpis primeiro.").font = DATA_FONT

        _auto_width(ws)

    # ── Sheet 6: Dados Brutos ──────────────────────────────────────────

    def _write_raw_data(self, wb) -> None:
        ws = wb.create_sheet("Dados Brutos")
        row = 1
        row = _write_title(ws, row, 1, "Dados Brutos — Todas as Bases")

        sheet_datasets = [
            ("faturamento_historico", "Faturamento Histórico"),
            ("contas_receber", "Contas a Receber"),
            ("contas_pagar", "Contas a Pagar"),
            ("estoque", "Estoque"),
            ("divida", "Dívida"),
            ("custo_vendas", "Custo de Vendas"),
        ]

        for key, label in sheet_datasets:
            df = self.datasets.get(key, pd.DataFrame())
            if df.empty:
                continue

            ws.cell(row=row, column=1, value=label).font = SUBTITLE_FONT
            row += 1

            if "data" in key.lower() or any("data" in c.lower() for c in df.columns):
                # Truncate to 30 rows for readability
                display_df = df.head(30).copy()
            else:
                display_df = df.copy()

            # Convert date columns
            for col in display_df.columns:
                if "data" in col.lower():
                    display_df[col] = pd.to_datetime(display_df[col])

            row = _write_dataframe(ws, display_df, start_row=row) + 1

            if len(df) > 30:
                ws.cell(row=row, column=1, value=f"({len(df)} registros no total — mostrando os primeiros 30)").font = Font(name="Calibri", size=8, color="999999", italic=True)
                row += 2

        _auto_width(ws)

    # ── Public API ─────────────────────────────────────────────────────

    def save(self, path: str) -> str:
        """Generate and save the full report to an .xlsx file."""
        from openpyxl import Workbook

        wb = Workbook()

        self._write_resumo(wb)
        self._write_forecast(wb)
        self._write_cashflow(wb)
        self._write_working_capital(wb)
        self._write_debt(wb)
        self._write_raw_data(wb)

        wb.save(path)
        return path


# ---------------------------------------------------------------------------
# Convenience function
# ---------------------------------------------------------------------------


def generate_report(
    datasets: dict[str, pd.DataFrame],
    forecast_df: Optional[pd.DataFrame] = None,
    cashflow_df: Optional[pd.DataFrame] = None,
    kpis: Optional[Any] = None,
    output_path: str = "data/outputs/relatorio_fpa.xlsx",
) -> str:
    """One-shot: generate the full FP&A report and save to disk.

    Returns the output path on success.
    """
    report = XLSXReport(
        datasets=datasets,
        forecast_df=forecast_df,
        cashflow_df=cashflow_df,
        kpis=kpis,
    )
    return report.save(output_path)


# ---------------------------------------------------------------------------
# Private helper
# ---------------------------------------------------------------------------


def _format_brl(value: float) -> str:
    """Internal BRL formatter (avoids import from utils if not available)."""
    try:
        from utils.formatting import format_brl
        return format_brl(value)
    except ImportError:
        if value is None:
            return "0,00"
        sign = "-" if value < 0 else ""
        value = abs(float(value))
        integer_part = int(value)
        decimal_part = int(round((value - integer_part) * 100))
        int_str = f"{integer_part:,}".replace(",", ".")
        return f"{sign}{int_str},{decimal_part:02d}"
